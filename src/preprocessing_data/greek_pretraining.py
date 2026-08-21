"""Add the Esteban and labelled Greek datasets to TrOCR pretraining.

The source datasets are added to both ``greek/pretraining`` and
``combined/pretraining``. Finetuning partitions are deliberately untouched.
Running the script repeatedly is safe: previously imported rows are replaced.
"""

from __future__ import annotations

import json
import os
import random
import shutil
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
PROCESSED_ROOT = REPO_ROOT / "data" / "processed"
ESTEBAN_ROOT = REPO_ROOT / "data" / "raw" / "greek_estaban"
LABELLED_ROOT = REPO_ROOT / "data" / "raw" / "greek_labelled_data"
TARGETS = (
    PROCESSED_ROOT / "greek" / "pretraining",
    PROCESSED_ROOT / "combined" / "pretraining",
)
SPLITS = ("train", "val", "test")
IMPORTED_PREFIXES = ("esteban__", "labelled__")
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".tif", ".tiff"}
SPLIT_SEED = 1111


@dataclass(frozen=True)
class ImportRow:
    split: str
    output_name: str
    text: str
    source_image: Path


def normalize_split(value: object) -> str:
    """Normalize source split names to the TrOCR manifest convention."""
    split = str(value).strip().lower()
    aliases = {"validation": "val", "valid": "val"}
    split = aliases.get(split, split)
    if split not in SPLITS:
        raise ValueError(f"Unsupported dataset split: {value!r}")
    return split


def read_esteban_rows(root: Path) -> list[ImportRow]:
    """Read Esteban image names, transcriptions, and splits from its workbook."""
    workbook_path = root / "gray_labels.xlsx"
    image_root = root / "all_bin"
    if not workbook_path.is_file() or not image_root.is_dir():
        raise FileNotFoundError(f"Incomplete Esteban dataset under {root}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook.active
    values = worksheet.iter_rows(values_only=True)
    try:
        header = next(values)
    except StopIteration as exc:
        raise ValueError(f"Empty Esteban workbook: {workbook_path}") from exc

    columns = {str(value).strip(): index for index, value in enumerate(header) if value}
    required = {"preproc_file_name", "label", "split"}
    missing = required - columns.keys()
    if missing:
        raise ValueError(f"Missing Esteban workbook columns: {sorted(missing)}")

    rows: list[ImportRow] = []
    seen_names: set[str] = set()
    for row_number, values_row in enumerate(values, start=2):
        image_value = values_row[columns["preproc_file_name"]]
        label_value = values_row[columns["label"]]
        split_value = values_row[columns["split"]]
        if image_value is None and label_value is None and split_value is None:
            continue
        if image_value is None or label_value is None or split_value is None:
            raise ValueError(f"Incomplete Esteban row {row_number}")

        source_name = str(image_value).strip()
        output_name = f"esteban__{source_name}"
        if output_name in seen_names:
            raise ValueError(f"Duplicate Esteban image in workbook: {source_name}")
        seen_names.add(output_name)

        source_image = image_root / source_name
        if not source_image.is_file():
            raise FileNotFoundError(f"Missing Esteban image: {source_image}")
        rows.append(
            ImportRow(
                split=normalize_split(split_value),
                output_name=output_name,
                text=str(label_value).strip(),
                source_image=source_image,
            )
        )
    workbook.close()
    return rows


def read_text(path: Path) -> str:
    """Read a ground-truth file, including legacy UTF-16 exports."""
    try:
        return path.read_text(encoding="utf-8").strip()
    except UnicodeDecodeError:
        return path.read_text(encoding="utf-16").strip()


def read_labelled_rows(root: Path) -> list[ImportRow]:
    """Read Kraken-style image and ``.gt.txt`` pairs from labelledData."""
    rows: list[ImportRow] = []
    for split in SPLITS:
        image_root = root / "images" / split
        label_root = root / "labels" / split
        if not image_root.is_dir() or not label_root.is_dir():
            raise FileNotFoundError(f"Incomplete labelledData split: {root}/{split}")

        image_paths = sorted(
            path
            for path in image_root.iterdir()
            if path.is_file() and path.suffix.lower() in IMAGE_SUFFIXES
        )
        for source_image in image_paths:
            label_path = label_root / f"{source_image.stem}.gt.txt"
            if not label_path.is_file():
                raise FileNotFoundError(f"Missing label for {source_image}: {label_path}")
            rows.append(
                ImportRow(
                    split=split,
                    output_name=f"labelled__{source_image.name}",
                    text=read_text(label_path),
                    source_image=source_image,
                )
            )
    return rows


def resplit_import_rows(rows: list[ImportRow], seed: int) -> list[ImportRow]:
    """Pool a source dataset and assign a deterministic 80/10/10 split."""
    shuffled = sorted(rows, key=lambda row: row.output_name)
    random.Random(seed).shuffle(shuffled)
    train_end = round(len(shuffled) * 0.8)
    val_end = round(len(shuffled) * 0.9)
    split_rows: list[ImportRow] = []
    for index, row in enumerate(shuffled):
        split = "train" if index < train_end else "val" if index < val_end else "test"
        split_rows.append(
            ImportRow(
                split=split,
                output_name=row.output_name,
                text=row.text,
                source_image=row.source_image,
            )
        )
    return split_rows


def parse_manifest(path: Path) -> list[tuple[str, str]]:
    """Read one TrOCR filename/transcription manifest."""
    rows: list[tuple[str, str]] = []
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        if not line:
            continue
        try:
            image_name, text = line.split("\t", 1)
        except ValueError as exc:
            raise ValueError(f"Invalid manifest row {path}:{line_number}") from exc
        rows.append((image_name, text))
    return rows


def hardlink_or_copy(source: Path, destination: Path) -> None:
    """Hardlink an image when possible, falling back to a copy."""
    try:
        os.link(source, destination)
    except OSError:
        shutil.copy2(source, destination)


def write_manifest(path: Path, rows: list[tuple[str, str]]) -> None:
    """Write one TrOCR manifest."""
    with path.open("w", encoding="utf-8", newline="\n") as output:
        for image_name, text in rows:
            output.write(f"{image_name}\t{text}\n")


def stage_partition(target: Path, imports: list[ImportRow]) -> tuple[Path, dict[str, int]]:
    """Build a validated replacement for one pretraining partition."""
    if not target.is_dir():
        raise FileNotFoundError(f"Missing TrOCR pretraining partition: {target}")

    staging = target.with_name(f".{target.parent.name}_pretraining_staging")
    if staging.exists():
        shutil.rmtree(staging)
    image_dir = staging / "image"
    image_dir.mkdir(parents=True)

    imported_by_split = {
        split: [row for row in imports if row.split == split] for split in SPLITS
    }
    summary: dict[str, int] = {}
    for split in SPLITS:
        manifest = target / f"gt_{split}.txt"
        if not manifest.is_file():
            raise FileNotFoundError(f"Missing TrOCR manifest: {manifest}")

        rows = [
            (image_name, text)
            for image_name, text in parse_manifest(manifest)
            if not image_name.startswith(IMPORTED_PREFIXES)
        ]
        seen_names = {image_name for image_name, _ in rows}
        if len(seen_names) != len(rows):
            raise ValueError(f"Duplicate existing rows in {manifest}")

        for image_name, _ in rows:
            source_image = target / "image" / image_name
            if not source_image.is_file():
                raise FileNotFoundError(f"Missing existing TrOCR image: {source_image}")
            hardlink_or_copy(source_image, image_dir / image_name)

        for import_row in imported_by_split[split]:
            if import_row.output_name in seen_names:
                raise ValueError(f"Duplicate imported image name: {import_row.output_name}")
            seen_names.add(import_row.output_name)
            hardlink_or_copy(
                import_row.source_image,
                image_dir / import_row.output_name,
            )
            rows.append((import_row.output_name, import_row.text))

        write_manifest(staging / f"gt_{split}.txt", rows)
        summary[split] = len(rows)
    return staging, summary


def replace_partitions(staged_targets: list[tuple[Path, Path]]) -> None:
    """Swap staged partitions into place, rolling both back on failure."""
    backups: list[tuple[Path, Path]] = []
    try:
        for target, staging in staged_targets:
            backup = target.with_name(f".{target.parent.name}_pretraining_backup")
            if backup.exists():
                shutil.rmtree(backup)
            target.rename(backup)
            backups.append((target, backup))
            staging.rename(target)
    except Exception:
        for target, backup in reversed(backups):
            if target.exists():
                shutil.rmtree(target)
            backup.rename(target)
        raise
    else:
        for _, backup in backups:
            shutil.rmtree(backup)


def main() -> None:
    """Import both Greek datasets into Greek and combined pretraining only."""
    esteban_rows = resplit_import_rows(
        read_esteban_rows(ESTEBAN_ROOT),
        SPLIT_SEED,
    )
    labelled_rows = resplit_import_rows(
        read_labelled_rows(LABELLED_ROOT),
        SPLIT_SEED + 1,
    )
    imports = esteban_rows + labelled_rows

    staged_targets: list[tuple[Path, Path]] = []
    target_summaries: dict[str, dict[str, int]] = {}
    try:
        for target in TARGETS:
            staging, summary = stage_partition(target, imports)
            staged_targets.append((target, staging))
            target_summaries[str(target.relative_to(PROCESSED_ROOT))] = summary
        replace_partitions(staged_targets)
    except Exception:
        for _, staging in staged_targets:
            if staging.exists():
                shutil.rmtree(staging)
        raise

    source_summary = {
        "esteban": {
            split: sum(row.split == split for row in esteban_rows) for split in SPLITS
        },
        "labelled": {
            split: sum(row.split == split for row in labelled_rows) for split in SPLITS
        },
    }
    print(
        json.dumps(
            {
                "sources": source_summary,
                "updated_pretraining_partitions": target_summaries,
                "finetuning_partitions_changed": False,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
